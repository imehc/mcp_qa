"""
Excel文档解析器
支持.xlsx和.xls格式的Excel文件解析
"""

import os
import logging
from typing import Dict, Any, List, Optional
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from langchain.text_splitter import RecursiveCharacterTextSplitter

from ..exceptions import ParsingError

logger = logging.getLogger(__name__)


class ExcelParser:
    """Excel文档解析器"""
    
    def __init__(self):
        """初始化Excel解析器"""
        self.supported_extensions = ['.xlsx', '.xls']
        self.text_splitter = RecursiveCharacterTextSplitter(
            chunk_size=1000,
            chunk_overlap=200,
            separators=["\n\n", "\n", "，", "。", "；", "：", ",", ".", ";", ":"]
        )
    
    def can_parse(self, file_path: str) -> bool:
        """检查是否可以解析该文件"""
        return any(file_path.lower().endswith(ext) for ext in self.supported_extensions)
    
    def parse(self, file_path: str, **kwargs) -> Dict[str, Any]:
        """
        解析Excel文件
        
        Args:
            file_path: Excel文件路径
            **kwargs: 其他参数
                - max_rows: 最大读取行数，默认None（无限制）
                - include_empty: 是否包含空行，默认False
                - sheet_names: 指定要解析的工作表名称列表，默认None（解析所有）
        
        Returns:
            解析结果字典
        """
        if not os.path.exists(file_path):
            raise ParsingError(f"文件不存在: {file_path}")
        
        if not self.can_parse(file_path):
            raise ParsingError(f"不支持的文件格式: {file_path}")
        
        try:
            max_rows = kwargs.get('max_rows', None)
            include_empty = kwargs.get('include_empty', False)
            sheet_names = kwargs.get('sheet_names', None)
            
            logger.info(f"开始解析Excel文件: {file_path}")
            
            # 使用pandas读取Excel文件
            if sheet_names:
                # 读取指定的工作表
                excel_data = pd.read_excel(
                    file_path, 
                    sheet_name=sheet_names,
                    nrows=max_rows,
                    keep_default_na=not include_empty
                )
            else:
                # 读取所有工作表
                excel_data = pd.read_excel(
                    file_path, 
                    sheet_name=None,  # 读取所有工作表
                    nrows=max_rows,
                    keep_default_na=not include_empty
                )
            
            # 如果只有一个工作表，pandas返回DataFrame，否则返回字典
            if isinstance(excel_data, pd.DataFrame):
                excel_data = {'Sheet1': excel_data}
            
            sheets_info = []
            all_text_content = []
            total_rows = 0
            total_cols = 0
            
            for sheet_name, df in excel_data.items():
                if df.empty:
                    continue
                
                # 处理NaN值
                df = df.fillna('')
                
                # 工作表信息
                sheet_rows, sheet_cols = df.shape
                total_rows += sheet_rows
                total_cols = max(total_cols, sheet_cols)
                
                # 将DataFrame转换为文本内容
                sheet_text = self._dataframe_to_text(df, sheet_name)
                all_text_content.append(sheet_text)
                
                # 获取列信息
                columns_info = []
                for idx, col in enumerate(df.columns):
                    col_letter = get_column_letter(idx + 1)
                    columns_info.append({
                        "index": idx,
                        "letter": col_letter,
                        "name": str(col),
                        "type": str(df[col].dtype),
                        "non_null_count": df[col].count()
                    })
                
                # 获取数据样本（前5行）
                sample_data = []
                for _, row in df.head(5).iterrows():
                    sample_data.append([str(cell) for cell in row.values])
                
                sheet_info = {
                    "name": sheet_name,
                    "rows": sheet_rows,
                    "columns": sheet_cols,
                    "columns_info": columns_info,
                    "sample_data": sample_data,
                    "text_content": sheet_text
                }
                sheets_info.append(sheet_info)
            
            # 合并所有文本内容
            full_text = "\n\n".join(all_text_content)
            
            # 文本分块
            chunks = self.text_splitter.split_text(full_text) if full_text.strip() else []
            
            # 使用openpyxl获取更详细的信息（如果是.xlsx文件）
            additional_info = {}
            if file_path.lower().endswith('.xlsx'):
                try:
                    additional_info = self._get_excel_metadata(file_path)
                except Exception as e:
                    logger.warning(f"获取Excel元数据失败: {e}")
            
            result = {
                "file_path": file_path,
                "file_type": "excel",
                "total_sheets": len(sheets_info),
                "total_rows": total_rows,
                "total_columns": total_cols,
                "sheets": sheets_info,
                "full_text": full_text,
                "chunks": chunks,
                "chunk_count": len(chunks),
                "metadata": additional_info
            }
            
            logger.info(f"Excel文件解析完成: {len(sheets_info)}个工作表, {total_rows}行数据")
            return result
            
        except Exception as e:
            logger.error(f"Excel文件解析失败: {file_path}, 错误: {e}")
            raise ParsingError(f"Excel解析失败: {str(e)}")
    
    def _dataframe_to_text(self, df: pd.DataFrame, sheet_name: str) -> str:
        """将DataFrame转换为文本格式"""
        lines = [f"工作表: {sheet_name}"]
        
        # 添加列标题
        headers = [str(col) for col in df.columns]
        lines.append("列标题: " + " | ".join(headers))
        lines.append("")
        
        # 添加数据行
        for idx, row in df.iterrows():
            row_data = [str(cell) if pd.notna(cell) else '' for cell in row.values]
            # 过滤空行
            if any(cell.strip() for cell in row_data):
                lines.append(" | ".join(row_data))
        
        return "\n".join(lines)
    
    def _get_excel_metadata(self, file_path: str) -> Dict[str, Any]:
        """获取Excel文件的元数据"""
        try:
            workbook = openpyxl.load_workbook(file_path, read_only=True)
            
            metadata = {
                "created": None,
                "modified": None,
                "creator": None,
                "title": None,
                "subject": None,
                "description": None,
                "keywords": None,
                "category": None,
                "comments": None
            }
            
            # 获取文档属性
            props = workbook.properties
            if props:
                metadata.update({
                    "created": str(props.created) if props.created else None,
                    "modified": str(props.modified) if props.modified else None,
                    "creator": props.creator,
                    "title": props.title,
                    "subject": props.subject,
                    "description": props.description,
                    "keywords": props.keywords,
                    "category": props.category,
                    "comments": props.comments
                })
            
            # 获取工作表名称
            metadata["sheet_names"] = workbook.sheetnames
            
            workbook.close()
            return metadata
            
        except Exception as e:
            logger.warning(f"获取Excel元数据失败: {e}")
            return {}
    
    def extract_table_structure(self, file_path: str, sheet_name: Optional[str] = None) -> Dict[str, Any]:
        """
        提取Excel文件的表格结构信息
        
        Args:
            file_path: Excel文件路径
            sheet_name: 工作表名称，如果为None则分析所有工作表
        
        Returns:
            表格结构信息
        """
        try:
            if sheet_name:
                df = pd.read_excel(file_path, sheet_name=sheet_name)
                sheets_data = {sheet_name: df}
            else:
                sheets_data = pd.read_excel(file_path, sheet_name=None)
                if isinstance(sheets_data, pd.DataFrame):
                    sheets_data = {'Sheet1': sheets_data}
            
            structures = {}
            
            for name, df in sheets_data.items():
                if df.empty:
                    continue
                
                # 分析数据类型
                dtype_info = {}
                for col in df.columns:
                    dtype_info[str(col)] = {
                        "dtype": str(df[col].dtype),
                        "null_count": df[col].isnull().sum(),
                        "unique_count": df[col].nunique(),
                        "sample_values": df[col].dropna().head(3).tolist()
                    }
                
                # 检测可能的表头行
                header_candidates = []
                for i in range(min(5, len(df))):
                    row = df.iloc[i]
                    if row.isnull().sum() < len(row) * 0.5:  # 非空值超过50%
                        header_candidates.append(i)
                
                structures[name] = {
                    "shape": df.shape,
                    "columns": list(df.columns),
                    "dtypes": dtype_info,
                    "header_candidates": header_candidates,
                    "has_numeric_data": any(df.select_dtypes(include=['number']).columns),
                    "has_datetime_data": any(df.select_dtypes(include=['datetime']).columns),
                    "memory_usage": df.memory_usage(deep=True).sum()
                }
            
            return {
                "file_path": file_path,
                "structures": structures,
                "total_sheets": len(structures)
            }
            
        except Exception as e:
            logger.error(f"提取Excel表格结构失败: {e}")
            raise ParsingError(f"提取表格结构失败: {str(e)}")


def create_excel_parser() -> ExcelParser:
    """创建Excel解析器实例"""
    return ExcelParser()


# 测试函数
def test_excel_parser():
    """测试Excel解析器"""
    parser = ExcelParser()
    
    # 测试文件类型检查
    assert parser.can_parse("test.xlsx")
    assert parser.can_parse("test.xls")
    assert not parser.can_parse("test.pdf")
    
    logger.info("Excel解析器测试通过")


if __name__ == "__main__":
    test_excel_parser()